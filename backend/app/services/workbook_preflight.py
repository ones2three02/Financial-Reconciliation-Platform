from datetime import date, datetime
from zipfile import BadZipFile

from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from backend.app.crud.field_mapping import get_mappings_by_source
from backend.app.domain.extraction_profiles import get_profile
from backend.app.schemas.preflight import PreflightResult
from backend.app.services.cleaner import clean_date
from backend.app.services.field_binding import FieldBindingError, resolve_field_bindings
from backend.app.services.workbook_io import WorkbookArchiveLimitError, load_data_workbook
from backend.app.services.workbook_rows import is_summary_row


MAX_FILE_SIZE = 50 * 1024 * 1024
MAX_SHEETS = 50
MAX_ROWS = 200_000
MAX_COLUMNS = 200


class PreflightValidationError(ValueError):
    """文件不能安全进入提取流程。"""


class TemplateMismatchError(PreflightValidationError):
    """工作簿结构与声明模板不一致。"""


class WorkbookLimitError(PreflightValidationError):
    """工作簿超过允许的资源上限。"""


def _parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    parsed = clean_date(value)
    return parsed.date() if isinstance(parsed, datetime) else parsed


def _clean_header(value: object) -> str:
    return "" if value is None else str(value).strip()


def preflight_workbook(
    content: bytes,
    profile_code: str,
    business_date: date,
    store_id: int | None,
    db: Session | None = None,
    password: str | None = None,
) -> PreflightResult:
    profile = get_profile(profile_code)
    if not content:
        raise PreflightValidationError("工作簿内容为空")
    if len(content) > MAX_FILE_SIZE:
        raise WorkbookLimitError(f"工作簿超过 {MAX_FILE_SIZE // 1024 // 1024}MB 限制")
    if profile.requires_store_id and store_id is None:
        raise PreflightValidationError("该模板必须指定归属标准门店")

    try:
        workbook = load_data_workbook(content, password=password)
    except WorkbookArchiveLimitError as exc:
        raise WorkbookLimitError(str(exc)) from exc
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise PreflightValidationError("无法读取工作簿，请确认文件格式有效") from exc

    try:
        if len(workbook.sheetnames) > MAX_SHEETS:
            raise WorkbookLimitError(f"工作簿工作表数量超过 {MAX_SHEETS} 个限制")

        sheet_name = next(
            (name for name in profile.sheet_names if name in workbook.sheetnames),
            None,
        )
        if sheet_name is None:
            expected = "、".join(profile.sheet_names)
            raise TemplateMismatchError(
                f"模板 {profile.code} 需要工作表: {expected}"
            )

        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
        if (max_row is not None and max_row > MAX_ROWS) or (max_col is not None and max_col > MAX_COLUMNS):
            raise WorkbookLimitError(
                f"工作表规模超过限制: 最大 {MAX_ROWS} 行、{MAX_COLUMNS} 列"
            )

        header_values = next(
            sheet.iter_rows(
                min_row=profile.header_row,
                max_row=profile.header_row,
                values_only=True,
            ),
            (),
        )
        headers = [_clean_header(value) for value in header_values]
        mappings = (
            get_mappings_by_source(
                db,
                data_source=profile.input_source,
                is_active_only=False,
            )
            if db is not None
            else []
        )
        try:
            bindings = resolve_field_bindings(profile, headers, mappings)
        except FieldBindingError as exc:
            raise TemplateMismatchError(str(exc)) from exc

        date_column = bindings[profile.date_field]
        date_index = headers.index(date_column)
        store_index = (
            headers.index(bindings[profile.store_field])
            if profile.store_field is not None
            else None
        )
        total_data_rows = 0
        matching_row_count = 0
        parsed_dates: list[date] = []
        detected_stores: set[str] = set()

        for excel_row_number, row in enumerate(
            sheet.iter_rows(min_row=profile.header_row + 1, values_only=True),
            start=profile.header_row + 1,
        ):
            if not any(value not in (None, "") for value in row):
                continue
            # 过滤交易日期为空的无效行
            if date_index < len(row):
                raw_date_val = row[date_index]
                if raw_date_val is None or str(raw_date_val).strip() == "":
                    continue
            content_row = {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(headers)
            }
            if is_summary_row(profile, content_row, bindings):
                continue
            total_data_rows += 1
            try:
                row_date = _parse_date(row[date_index])
            except (ValueError, TypeError, IndexError) as exc:
                raise PreflightValidationError(
                    f"模板 {profile.code} 工作表 {sheet_name} 第 {excel_row_number} 行的"
                    f"{date_column}无法解析"
                ) from exc
            parsed_dates.append(row_date)
            if row_date == business_date:
                matching_row_count += 1
            if store_index is not None and store_index < len(row):
                raw_store = row[store_index]
                if raw_store not in (None, ""):
                    detected_stores.add(str(raw_store).strip())

        return PreflightResult(
            profile_code=profile.code,
            profile_version=profile.version,
            sheet_name=sheet_name,
            business_date=business_date,
            store_id=store_id,
            output_sources=list(profile.output_sources),
            total_data_rows=total_data_rows,
            matching_row_count=matching_row_count,
            date_range_start=min(parsed_dates) if parsed_dates else None,
            date_range_end=max(parsed_dates) if parsed_dates else None,
            detected_store_names=sorted(detected_stores),
        )
    finally:
        workbook.close()
