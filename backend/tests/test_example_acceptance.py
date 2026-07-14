from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from backend.app.models.store import Store, StoreAlias
from backend.app.services.batch_service import get_or_create_batch
from backend.app.services.cleaner import clean_date
from backend.app.services.import_pipeline import ImportWorkbookCommand, import_workbook
from backend.app.services.reconciliation_service import reconcile_batch
from backend.app.services.store_resolution import confirm_alias
from backend.app.services.workbook_io import load_data_workbook


BUSINESS_DATE = date(2026, 7, 10)
EXAMPLE_DIR = Path(__file__).resolve().parents[2] / "example"
MINYUAN_ALIASES = {
    "douyin": ("山道曙光游泳馆(曙光商贸城民族大道店)",),
    "meituan": ("武汉 : 山道健身游泳舞蹈(曙光店)",),
    "tonglian": (
        "山道健身－民院健身房",
        "山道健身－民院游泳馆",
    ),
}
PROFILE_CONFIG = {
    "douyin": ("抖音收入.xlsx", "核销明细", 1, "核销时间", "核销门店", "douyin_v1"),
    "meituan": ("美团收入.xlsx", "收益明细表", 1, "验券/退款/", "消费门店", "meituan_v1"),
    "tonglian": ("通联好老板.xlsx", "sheet1", 2, "统计日期", "门店名", "tonglian_v1"),
}


def scoped_channel_workbook(source: str) -> bytes:
    filename, sheet_name, header_row, date_column, store_column, _ = PROFILE_CONFIG[source]
    source_workbook = load_data_workbook((EXAMPLE_DIR / filename).read_bytes())
    try:
        source_sheet = source_workbook[sheet_name]
        header_values = next(
            source_sheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            )
        )
        headers = [str(value).strip() if value is not None else "" for value in header_values]
        date_index = headers.index(date_column)
        store_index = headers.index(store_column)
        matching_rows = []
        for row in source_sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue
            raw_store = str(row[store_index] or "").strip()
            if raw_store not in MINYUAN_ALIASES[source]:
                continue
            if clean_date(row[date_index]) != BUSINESS_DATE:
                continue
            matching_rows.append(list(row))
    finally:
        source_workbook.close()

    scoped_workbook = Workbook()
    target_sheet = scoped_workbook.active
    target_sheet.title = sheet_name
    for _ in range(1, header_row):
        target_sheet.append(["报表说明"])
    target_sheet.append(list(header_values))
    for row in matching_rows:
        target_sheet.append(row)
    output = BytesIO()
    scoped_workbook.save(output)
    return output.getvalue()


def test_minyuan_2026_07_10_reconciles_to_zero(db_session):
    store = Store(name="民院店", code="MD010", is_active=True)
    db_session.add(store)
    batch = get_or_create_batch(db_session, BUSINESS_DATE, actor="admin")
    db_session.commit()

    import_workbook(
        db_session,
        ImportWorkbookCommand(
            batch_id=batch.id,
            filename="7月民院财务表.xlsx",
            content=(EXAMPLE_DIR / "7月民院财务表.xlsx").read_bytes(),
            profile_code="store_finance_v1",
            store_id=store.id,
            actor="admin",
        ),
    )
    for source, (filename, _, _, _, _, profile_code) in PROFILE_CONFIG.items():
        import_workbook(
            db_session,
            ImportWorkbookCommand(
                batch_id=batch.id,
                filename=filename,
                content=scoped_channel_workbook(source),
                profile_code=profile_code,
                store_id=None,
                actor="admin",
            ),
        )

    before_confirmation = reconcile_batch(db_session, batch.id)[0]
    assert before_confirmation.status == "incomplete"

    for source, alias_names in MINYUAN_ALIASES.items():
        for alias_name in alias_names:
            alias = db_session.query(StoreAlias).filter_by(
                source_code=source,
                alias_name=alias_name,
            ).one()
            confirm_alias(
                db_session,
                alias_id=alias.id,
                store_id=store.id,
                actor="admin",
            )
    db_session.commit()

    result = reconcile_batch(db_session, batch.id)[0]
    assert result.tonglian_amount == Decimal("20480.00")
    assert result.meituan_amount == Decimal("9.90")
    assert result.douyin_amount == Decimal("296.00")
    assert result.sales_amount == Decimal("20785.90")
    assert result.cash_amount == Decimal("0.00")
    assert result.difference == Decimal("0.00")
    assert result.status == "consistent"
