from datetime import date, datetime
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from backend.app.models.field_mapping import FieldMapping
from backend.app.services.workbook_preflight import (
    PreflightValidationError,
    TemplateMismatchError,
    WorkbookLimitError,
    preflight_workbook,
)


def workbook_bytes(sheet_name: str, headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def tonglian_workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["交易统计汇总"])
    sheet.append(["统计日期", "门店名", "成功交易金额"])
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_tonglian_profile_ignores_explicit_summary_footer():
    result = preflight_workbook(
        tonglian_workbook_bytes([
            ["2026-07-10", "民院店原始名称", 100],
            ["汇总", None, 100],
        ]),
        profile_code="tonglian_v1",
        business_date=date(2026, 7, 10),
        store_id=None,
    )
    assert result.total_data_rows == 1
    assert result.matching_row_count == 1


def test_tonglian_profile_uses_configured_store_alias(db_session):
    db_session.add_all([
        FieldMapping(data_source="tonglian", target_field="trade_date", source_column="统计日期", is_active=True),
        FieldMapping(data_source="tonglian", target_field="store_name", source_column="门店名称", is_active=True),
        FieldMapping(data_source="tonglian", target_field="amount", source_column="成功交易金额", is_active=True),
    ])
    db_session.commit()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["交易统计汇总"])
    sheet.append(["统计日期", "门店名称", "成功交易金额"])
    sheet.append(["2026-07-10", "民院店", 100])
    output = BytesIO()
    workbook.save(output)

    result = preflight_workbook(
        output.getvalue(),
        profile_code="tonglian_v1",
        business_date=date(2026, 7, 10),
        store_id=None,
        db=db_session,
    )

    assert result.detected_store_names == ["民院店"]


def test_tonglian_profile_rejects_two_configured_aliases_in_same_file(db_session):
    db_session.add_all([
        FieldMapping(data_source="tonglian", target_field="trade_date", source_column="统计日期", is_active=True),
        FieldMapping(data_source="tonglian", target_field="store_name", source_column="门店名", is_active=True),
        FieldMapping(data_source="tonglian", target_field="store_name", source_column="门店名称", is_active=True),
        FieldMapping(data_source="tonglian", target_field="amount", source_column="成功交易金额", is_active=True),
    ])
    db_session.commit()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["交易统计汇总"])
    sheet.append(["统计日期", "门店名", "门店名称", "成功交易金额"])
    sheet.append(["2026-07-10", "民院店", "民院店", 100])
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(TemplateMismatchError, match="同时命中多个列"):
        preflight_workbook(
            output.getvalue(),
            profile_code="tonglian_v1",
            business_date=date(2026, 7, 10),
            store_id=None,
            db=db_session,
        )


@pytest.mark.parametrize("store_value", [0, False], ids=["zero", "false"])
def test_tonglian_profile_reports_sheet_and_row_for_bad_business_date(store_value):
    with pytest.raises(PreflightValidationError) as exc_info:
        preflight_workbook(
            tonglian_workbook_bytes([
                ["2026-07-10", "正常门店", 100],
                ["汇总", store_value, 50],
            ]),
            profile_code="tonglian_v1",
            business_date=date(2026, 7, 10),
            store_id=None,
        )
    message = str(exc_info.value)
    assert message == "模板 tonglian_v1 工作表 sheet1 第 4 行的统计日期无法解析"
    assert "汇总" not in message


def test_finance_profile_requires_exact_sheet_and_store():
    content = workbook_bytes(
        "收入流水表",
        ["日期", "付款方式", "金额"],
        [
            [datetime(2026, 7, 10), "微信", 100],
            [datetime(2026, 7, 11), "现金", 20],
        ],
    )

    result = preflight_workbook(
        content,
        profile_code="store_finance_v1",
        business_date=date(2026, 7, 10),
        store_id=10,
    )

    assert result.sheet_name == "收入流水表"
    assert result.output_sources == ["sales", "cash"]
    assert result.matching_row_count == 1
    assert result.date_range_start == date(2026, 7, 10)
    assert result.date_range_end == date(2026, 7, 11)


def test_finance_profile_rejects_missing_file_store():
    content = workbook_bytes(
        "收入流水表",
        ["日期", "付款方式", "金额"],
        [[datetime(2026, 7, 10), "微信", 100]],
    )

    with pytest.raises(PreflightValidationError, match="标准门店"):
        preflight_workbook(
            content,
            profile_code="store_finance_v1",
            business_date=date(2026, 7, 10),
            store_id=None,
        )


def test_meituan_profile_rejects_missing_marketing_column():
    content = workbook_bytes(
        "收益明细表",
        ["验券/退款/", "消费门店", "总收入（元）"],
        [[datetime(2026, 7, 10), "脱敏门店", 100]],
    )

    with pytest.raises(TemplateMismatchError, match="商家营销费用（元）"):
        preflight_workbook(
            content,
            profile_code="meituan_v1",
            business_date=date(2026, 7, 10),
            store_id=None,
        )


def test_profile_does_not_guess_another_sheet():
    content = workbook_bytes(
        "其他明细",
        ["核销时间", "核销门店", "订单实收"],
        [[datetime(2026, 7, 10), "脱敏门店", 100]],
    )

    with pytest.raises(TemplateMismatchError, match="核销明细"):
        preflight_workbook(
            content,
            profile_code="douyin_v1",
            business_date=date(2026, 7, 10),
            store_id=None,
        )


def test_template_error_does_not_include_data_rows():
    content = workbook_bytes(
        "收益明细表",
        ["验券/退款/", "消费门店", "总收入（元）"],
        [[datetime(2026, 7, 10), "含敏感姓名张三的门店", 100]],
    )

    with pytest.raises(TemplateMismatchError) as exc_info:
        preflight_workbook(
            content,
            profile_code="meituan_v1",
            business_date=date(2026, 7, 10),
            store_id=None,
        )

    assert "张三" not in str(exc_info.value)


def test_high_compression_ratio_is_rejected_before_workbook_parse():
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"A" * 1_000_000)

    with pytest.raises(WorkbookLimitError, match="压缩"):
        preflight_workbook(
            output.getvalue(),
            profile_code="douyin_v1",
            business_date=date(2026, 7, 10),
            store_id=None,
        )
